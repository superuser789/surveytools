
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM, BORDER_THICK
import pandas as pd
from natsort import natsorted
import os, sys
from time import sleep
from tqdm import tqdm
import warnings
warnings.filterwarnings('ignore')



verbose=False
#verbose=True
progressmenu=1
startRow=2
plotOne=['-', 0] # SubPlot to Take as 1 
plotsDF = pd.DataFrame(columns=['plotNo','subplot','sno','status']) # DataFrame
n=0 # DataFrame Index



def getNumber(cstr): # Get Number at the starting of s string
    num=''
    for c in cstr:
      if c.isdigit():
          num+=c
      else:
          break
    return int(num)

def splitAlpnaNum(sps):
    divlist=[]; tstr=''
    for sp in sps:
        if not sp.isdigit(): # There is alphabet
            divlist.append(tstr) # Add previous numbers
            divlist.append(sp) # Add Current alphabet
            tstr='' # Reset the no.
        else:
            tstr+=sp
    divlist.append(tstr)
    return [ i for i in divlist if i ]
    #import re; return [ i for i in re.split('(\d+)', sps) if i ]

def clear():
    from IPython import get_ipython
    get_ipython().run_line_magic('clear','')
    os.system('cls' if os.name=='nt' else 'clear')


# Gives the Passed Arguments
def getArg(usagestr='', customarg=''): # Gives arguments & Takes Single Custom Arg. optionally
    args = sys.argv[1:] # Exclude Filename from obtained Arguments list
    if len(args) < 1:
        print(usagestr)
        return {}
    else:
        if len(args) > 2:
            if customarg:
                if args[1] == customarg or args[1] == '--'+customarg or args[1] == '-'+customarg :
                    return { 'main': args[0], customarg : args[2] }    
        else:
            return { 'main': args[0] }


pval='' ; pltseq=[]
def findMissingPlots(subp): # Takes Subplots of a Plot
    ### Subplots into Dictionary of subplots stored in subplt list
    subplt=[]
    for sps in subp:
        tmpdict=currdict={}
        #spsNo=str(getNumber(sps))
        for sp in splitAlpnaNum(sps) : # Get First 1,2or3 digit no. from subplot
            currdict[sp] = {}
            currdict=currdict[sp]
        subplt.append(tmpdict)

    ### Adding Missing Plots using subplt list
    allplt=[]
    global pval, pltseq
    '''
    def getseq(plt, prevplt):
       extra=[]
       startPlt=65 if ord(plt) > 64 else 49 # startRange for Alphabet & Number
       #LastPlt = 1 if len(prevplt) > 1 else 0
       #if len(prevplt) > 1 :
       extra=['###'+prevplt]
       return [ str(prevplt)+chr(i) for i in range(startPlt, ord(plt)+1)] + extra
    '''
    def getseq(plt, prevplt):
       #print(plt,':', prevplt)
       extra=[]
       #if len(prevplt+plt) == 2 : # Converts xA into x
       if plt=='A':
           plt='B'
       if plt=='1':
           plt='2'
               
       if plt.isdigit():  # startRange & stopRange for Alphabet & Number
           startPlt=1
           stopPlt=int(plt)
           Intflag=True
       else:
           startPlt=65
           stopPlt=ord(plt)
           Intflag=False
       #startPlt=65 if ord(plt) > 64 else 49 # startRange for Alphabet & Number
       # if len(prevplt+plt) == 2 : # Converts xA into x
       #     if plt=='A':
       #         return [ prevplt, '###'+prevplt+plt ]
       extra=['###'+prevplt]
       return [ str(prevplt)+ ( str(i) if Intflag else chr(i) ) for i in range(startPlt, stopPlt+1)] + extra


  
    def getall(dplt):
        global pval, pltseq
        for k,v in dplt.items():
            #print(k, ',', pval)
            if pval:
                pltseq+=getseq(k, pval)
            pval+=str(k)
            if isinstance(v, dict):
                getall(v)
    
    for sp in subplt: # subplt dict in subplt list
        pval='' ; pltseq=[]
        getall(sp) ; #print(pltseq)
        allplt+=pltseq

    return allplt
    

MissedSinglePlots={}
def getsubplots( subplot, pno=0, addPlot=0): # Gives All Possible subplots of a Plot
    '''
    Pass the List having Subplots in random order to get the Sequenced Subplots
    '''
    global n, plotsDF, plotOne, MissedSinglePlots
    if not addPlot:
        if not pno:
            plotsDF = pd.DataFrame(columns=['plotNo','subplot','sno','status']); n=0; MissedSinglePlots={}
            
            subplot=[int(i) if str(i).isdigit() else str(i) for i in subplot.split(',')] if isinstance(subplot, str) else subplot
            subplot=[ 1 if i in plotOne else i for i in subplot ]


        mainRange = max(set([ i if isinstance(i,int) else getNumber(i)  for i in subplot ])) # Get Max. Main Subplots
        SinglePlots = [ str(i) for i in subplot if isinstance(i,int) ] # Single Plots into string
        GivenSubPlots=[ i for i in subplot if isinstance(i,str) ] # All SubPlots are string type
        MissedPlots = list(set(findMissingPlots(GivenSubPlots)))
        GivenSingleSubPlots = set([str(getNumber(i)) for i in MissedPlots if i[:3] != '###'])
        if verbose:
            print(MissedPlots)
        MissedPlot=MissedPlots.copy()
        ### Remove Extra Plots &  Only single occurrence due to set
        for plt in MissedPlot:
            if '###' == plt[:3]:
                MissedPlots.remove(plt)
                if plt[3:] in MissedPlot:
                    MissedPlots.remove(plt[3:])
                if plt[3:].isdigit() and int(plt[3:]) in MissedPlot: # In case of Integer Plot No.
                    MissedPlots.remove(int(plt[3:]))
        
        [ SinglePlots.remove(i)  for i in GivenSingleSubPlots if i in SinglePlots ] # Remove SinglePlots if it is in GivenSubPlots 
        
    
        #AddedPlots = AllPlots - set(GivenSubPlots)
        #AllPlots = sorted( set(SinglePlots) | set(MissedPlots) , key=lambda x: int(''.join(filter(str.isdigit, x))) ) # Combine & Sort
        AllPlots = natsorted( set(SinglePlots) | set(MissedPlots) ) # Combine & Sort
        if verbose:
            print(AllPlots)    
        ## Insert Missing Single Plot if any
        singleplt = list({ getNumber(i) for i in AllPlots }) # All SubPlot Nos. 
        isingleplt = [ i for i in range(1, mainRange+1) ]
        
        if singleplt != isingleplt :
            AddedIntPlots=[ str(i)+'++' for i in isingleplt if i not in singleplt ]
            AllPlots = natsorted( set(SinglePlots+AddedIntPlots) | set(MissedPlots) ) # Combine & Sort    
            if verbose:
                print('\n  Missing In Plot :',pno, ' - ',singleplt, '  Added: ',[i[:-2] for i in AddedIntPlots])
            MissedSinglePlots[pno] = [ singleplt, AddedIntPlots ]
        

        for plt in AllPlots :
            n+=1
            plotsDF.at[n, 'plotNo'] = pno # Add Plot No.
            plotsDF.at[n,'subplot'] = plt
            plotsDF.at[n,'sno'] = n
            
            if plt in SinglePlots : # If Single Plot or Integer
                plotsDF.at[n, 'status'] = 'single'                    
            elif plt in GivenSubPlots:
                plotsDF.at[n, 'status'] = 'division'
            elif '++' == plt[-2:]:
                plotsDF.at[n,'subplot'] = plt[:-2] # overwrite in case of ++ single Plots
                plotsDF.at[n, 'status'] = 'ADDED'
            else:                
                plotsDF.at[n, 'status'] = 'ADDED'                
    
        if verbose:
            print(plotsDF.tail())
    else: # Plot is not Added
        n+=1
        plotsDF.at[n, 'plotNo'] = pno # Add Plot No.
        plotsDF.at[n,'subplot'] = subplot
        plotsDF.at[n,'sno'] = n
        plotsDF.at[n, 'status'] = 'ADDED' 
            
    return  [ i for i in plotsDF['subplot'] ]


def getFilename(FILE_NAME, noext=0):
    FileFormat='.xls'
    if FileFormat in FILE_NAME:
        FileName, extension = str(FILE_NAME).split(FileFormat)
        return FileName if noext else FileName+FileFormat+extension
    else:
        return FILE_NAME if noext else FILE_NAME+'.xlsx'

  
def createEXCEL(filename, checkfile=0 , autoOpen=1, selectedPlots=0):
    global plotsDF, startRow, FileFullPath
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.title =  'Final'
      
    START_COL='A'
    START_ROW='2'
      
    bold24Font = Font(size=11, bold=True)  # name='Times New Roman'
    bold26Font = Font(size=12, bold=True)  # name='Times New Roman'
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    leftalignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    titleColourTOP='E0E0E0'
    AddedPlotColour='CCFF90'
    
    thick_border = Border(
       left=Side(border_style=BORDER_MEDIUM, color='00000000'),
       right=Side(border_style=BORDER_MEDIUM, color='00000000'),
       top=Side(border_style=BORDER_MEDIUM, color='00000000'),
       bottom=Side(border_style=BORDER_MEDIUM, color='00000000') )
    thin_border = Border(
       left=Side(border_style=BORDER_THIN, color='00000000'),
       right=Side(border_style=BORDER_THIN, color='00000000'),
       top=Side(border_style=BORDER_THIN, color='00000000'),
       bottom=Side(border_style=BORDER_THIN, color='00000000') )
       
    side_border = Border( right=Side(border_style=BORDER_MEDIUM, color='00000000') ) 
    top_border = Border( top=Side(border_style=BORDER_THIN, color='00000000') )
    
    for i in range(3):
        sheet[chr(ord(START_COL) + i)+'1'].font = bold24Font
        sheet[chr(ord(START_COL) + i)+'1'].alignment = alignment  
        sheet[chr(ord(START_COL) + i)+'1'].fill = PatternFill(fgColor=titleColourTOP, fill_type = 'solid')  
        sheet[chr(ord(START_COL) + i)+'1'].border = thick_border  

        
    ## Title Names
    sheet[chr(ord(START_COL) ) +str(1)] = 'Survey No.'
    sheet[chr(ord(START_COL)+1 )+str(1)] = 'Sub-Division'
    sheet[chr(ord(START_COL)+2 )+str(1)] = 'S.No.'

    nextPlotNo=1
    plotInfo={}
    for i in range(len(plotsDF.index)):
        plotInfo.setdefault(plotsDF.iloc[i,0],0)
        plotInfo[plotsDF.iloc[i,0]]+=1
        
        sheet[chr(ord(START_COL) )+str(int(START_ROW)+i)] = plotsDF.iloc[i,0] # plot No.
        sheet[chr(ord(START_COL)+1)+str(int(START_ROW)+i)]= int(plotsDF.iloc[i,1]) if plotsDF.iloc[i,1].isdigit() else plotsDF.iloc[i,1] # SubPlot
        sheet[chr(ord(START_COL)+2)+str(int(START_ROW)+i)]= plotsDF.iloc[i,2] # S.No.


        ## Add Line on Plot No. Change
        if nextPlotNo != plotsDF.iloc[i,0]:
            for j in range(3):
                sheet[chr(ord(START_COL)+j)+str(int(START_ROW)+i)].border = top_border
        nextPlotNo=plotsDF.iloc[i,0]
        
        ### Formatting for Newly Added Subplots
        if plotsDF.iloc[i,3] == 'ADDED' :
            for j in range(3):
                sheet[chr(ord(START_COL)+j)+str(int(START_ROW)+i)].fill = PatternFill(fgColor=AddedPlotColour, fill_type = 'solid')
                sheet[chr(ord(START_COL)+j)+str(int(START_ROW)+i)].border = thin_border
        
        
        ## Cell Formatting
        sheet[chr(ord(START_COL))+str(int(START_ROW)+i)].alignment = alignment
        sheet[chr(ord(START_COL)+1)+str(int(START_ROW)+i)].alignment = leftalignment
        sheet[chr(ord(START_COL)+2)+str(int(START_ROW)+i)].alignment = alignment

        
    if selectedPlots:
        sheet.title = 'Selected'
        splots=list(selectedPlots.keys())
        selectedPlots=[i for l in [selectedPlots[i] for i in selectedPlots] for i in l ]
        for i in range(len(plotsDF.index)):
            if plotsDF.iloc[i,2] in selectedPlots:
                pltcolor = '03A9F4' if sheet[chr(ord(START_COL)+2)+str(int(START_ROW)+i)].value in splots else 'B3E5FC'
                sheet[chr(ord(START_COL))+str(int(START_ROW)+i)].fill = PatternFill(fgColor=pltcolor, fill_type = 'solid') # 81D4FA
                
                sheet[chr(ord(START_COL)+1)+str(int(START_ROW)+i)].fill = PatternFill(fgColor=pltcolor, fill_type = 'solid')
                sheet[chr(ord(START_COL))+str(int(START_ROW)+i)].border = thin_border
                sheet[chr(ord(START_COL)+1)+str(int(START_ROW)+i)].border = thin_border

    else:
        sortedPlots = sorted(plotInfo, key=plotInfo.get, reverse=True) # Based on Subplots count
        print('\n\n     Plot No. | Divisions | Probability/%')
        for i,plt in enumerate(sortedPlots[:10]):
            print(' '+'{:>2}'.format(i+1)+'. '+'{:>5}'.format(plt)+'  '+'{:>9}'.format(plotInfo[plt])+'        '+format((plotInfo[plt]/len(plotsDF.index))*100, '.2f')+' %' )
            #print(' '+'{:>2}'.format(i+1)+'. '+'{:>3}'.format(plt)+'  '+'{:>3}'.format(plotInfo[plt])+'  '+format((plotInfo[plt]/len(plotsDF.index))*100, '.2f')+' %' )
        if checkfile:
            wbck = openpyxl.load_workbook(getFilename(checkfile))
            sheetck=wbck[wbck.sheetnames[-1]]
            sheetck=wbck.active
            
            ROW_MAX=sheetck.max_row
            
            chkplots={} # Dictionary to store Plot no. as key & its subplots as value in list form
            for i in range(startRow,ROW_MAX+1):
                #print(pno,' : ',spno)
                pno=sheetck['A'+str(i)].value # plot no.
                if pno:
                    spno = sheetck['B'+str(i)].value
                    sno = sheetck['C'+str(i)].value
                    chkplots[sno] = spno
            cn=int(START_ROW)
            for sno in chkplots:
                sheet['D'+str(cn)] = sno # SNo.
                sheet['E'+str(cn)] = chkplots[sno] # SNo.
                sheet['D'+str(cn)].alignment = alignment; sheet['E'+str(cn)].alignment = alignment
                cn+=1

    wb.save(FileFullPath+filename)  
            
    if autoOpen :
        sleep(4)
        os.system('start EXCEL.EXE "'+FileFullPath+filename+'"')


def selectplots(plotsDF):
    stmgrp=''
    if isinstance(plotsDF,str):
        tplot,stmgrp=tuple(plotsDF.split(',')) if ',' in plotsDF else (plotsDF,'')
    elif isinstance(plotsDF,list):
        tplot,stmgrp=(plotsDF[0],str(plotsDF[1])) if len(plotsDF) == 2 else (plotsDF[0],'')
    else:   
        tplot=len(plotsDF.index)
    tplot=int(tplot)
    def SelectedPlots(i):
        startRange=(i-i%10  if 0<i%10<6 else  i-i%5) if i%10 else i-5
        return [ startRange+i for i in range(1,6) ]

    rtplot=tplot+abs(tplot%5-5)
    splotinterval=int(rtplot/4)
    stmflag=True; grpflag=True
    print('\n\n')
    if not stmgrp:
        while(stmflag):
            stm=input('\n  - Enter Stratum : ')
            stmflag=False if stm.isdigit() else True
        while(grpflag):
            grp=input('  - Enter Order of Selection : ')
            grpflag=False if grp.isdigit() else True
        stmgrp=stm+grp
        
    if __name__ == '__main__' :
        RandomTable = pd.read_pickle('RandomTable.pkl')
    else:
        RandomTable = pd.read_pickle(os.path.dirname(os.path.realpath(__file__))+('\\' if os.name=='nt' else '/')+'RandomTable.pkl')
    if int(stmgrp) <=104 :
        for num in RandomTable[int(stmgrp)]:
            if int(num) <= tplot:
                RNUM=int(num)
                break
    else:
        print('\n\n\n     Random Column Out of Range : ',stmgrp,'\n\n\n')
        sys.exit()

    splots = [ RNUM+splotinterval*i  if RNUM+splotinterval*i <= tplot else (RNUM+splotinterval*i)-rtplot  for i in range(4) ] # Four Randomly Selected Plot SNo.

    print('\n\n   Random Number :',RNUM,'\n\n    H :',tplot," H' :",rtplot,'\n   Plot Interval : ',splotinterval)
    selectedplots={}
    if isinstance(plotsDF,pd.core.frame.DataFrame):
        sleep(2)
        print('\n\n    SELECTED PLOTS :: ')
        print('\n         SNo.     Plot - Subplot\n')
        n=1
        for plt in splots:
            selectedplots[plt]=SelectedPlots(plt)
            for sno in selectedplots[plt]:
                pno=plotsDF[plotsDF['sno']==sno]['plotNo'].values[0]
                subplt=plotsDF[plotsDF['sno']==sno]['subplot'].values[0]
                if sno==plt:
                    print(' [','{:>2}'.format(n)+'.','{:>5}'.format(sno),' ->  ',pno,'-',subplt,' ]')
                else:
                    print('  ','{:>2}'.format(n)+'.','{:>5}'.format(sno),' ->  ',pno,'-',subplt)
                print() if not n%5 else ''
                n+=1
    
        return selectedplots
    else:
        spltstr=[]; n=0
        for plt in splots:
            selectedplots[plt]=SelectedPlots(plt)
            for sno in selectedplots[plt]:
                n+=1
                if sno==plt:
                    spltstr.append('  {:>2}'.format(n)+'. [{:>5}'.format(sno)+']')
                else:
                    spltstr.append('  {:>2}'.format(n)+'.  {:>5}'.format(sno))
                spltstr.append('\n') if not n%5 else ''
        return { 'param':[ RNUM,rtplot,splotinterval ] , 'splots': spltstr }

        

FileFullPath=''
def agriplot(FILE_NAME, checkfile=''):
    '''
    Generates Excel file from Plot data file.
    Optionally, Pass the Excel file having manually done Plot sequences in the parameter 'check=' to check with the generated sequences.
    '''
    #print(vars(sys.modules[__name__])['__package__'])
    global FileFullPath
    FILENAME=getFilename(FILE_NAME)
    FileFullPath=os.path.abspath(FILENAME).replace(FILENAME,'')
    wb = openpyxl.load_workbook(getFilename(FILENAME))
    sheet=wb[wb.sheetnames[-1]]
    sheet=wb.active
    global startRow
    ROW_MAX=sheet.max_row
    
    perday=1000
    progressTime=18
    global plotOne
    #ROW_MAX=232
    plots={} # Dictionary to store Plot no. as key & its subplots as value in list form
    for i in range(startRow,ROW_MAX+1):
        #print(pno,' : ',spno)
        pno=sheet['A'+str(i)].value # plot no.
        if pno:
            #spno=sheet['B'+str(i)].value # its subplot
            spno = sheet['B'+str(i)].value if sheet['B'+str(i)].value not in plotOne else 1 # its subplot
            plots.setdefault(pno,[])
            plots[pno].append(spno)


    addedPlots=[]
    pn=0; ln=0
    Tplot=sum([ len(plots[pno]) if plots.get(pno) else 1  for pno in range(1,list(plots.keys())[-1]+1) ]) ; ln=0
    LastPlot=list(plots.keys())[-1] # For  Updatebar increment
    progressIncT=progressTime/Tplot # Updatebar wait
    with tqdm(total=100,bar_format='{desc}{percentage:3.1f}% | {bar}  [{elapsed}<{remaining}]' ) as pbar:    
        for plotNo in range(1,list(plots.keys())[-1]+1):
            if plots.get(plotNo):
                pn+=1
                ln+=len(plots[plotNo]) # Show Original Lines for Days calculation
                '''
                    clear()
                    progress = ('.  ' if pn%2 else ' . ' if pn%3 else '  .' )
                    
                    print('\n\n\n      Plot:','{:>3}'.format(plotNo),progress, '{:>5}'.format(int((pn/len(plots))*100)) ,'%  ')
                    print('\n      Days: ', ( 1 if ln <= perday else int(ln/perday)+1), '  |  Lines:','{:>5}'.format(ln) )
                '''

                subplots=getsubplots(plots[plotNo], plotNo)
                #uptL=(len(subplots)/Tplot)*100
                
            else:
                subplots=getsubplots('1' , plotNo, addPlot=plotNo)
                #ln+=1 # Uncomment to show Added Lines
                addedPlots.append(plotNo)
            if progressmenu:
                pbar.set_description("  Days:"+str( 1 if ln <= perday else int(ln/perday)+1)+' | Lines:'+str(ln)+' | Plots:'+str(plotNo)+' ')
                pbar.update(100/LastPlot)
                sleep(progressIncT)
                #sleep(0.005/uptL)


    print('\n\n\n')
    if MissedSinglePlots:
        for mplt in MissedSinglePlots:
            print('\n  Missing In Plot:',mplt, ' - ',MissedSinglePlots[mplt][0], '  Added: ',[i[:-2] for i in MissedSinglePlots[mplt][1]])
    if addedPlots:
        print('\n\n  Added Single Plots:',addedPlots)
        
    TotalAdded=plotsDF[plotsDF['status']=='ADDED']['status']
    if not TotalAdded.empty:
        TotalAddedPlots=int(TotalAdded.value_counts())
        print('\n\n  TOTAL Plots/SubPlots[Lines] Added: ',TotalAddedPlots,'\n')
        #print('  TOTAL Plots/SubPlots : ',len(plotsDF.index),' = ',ln,'+',TotalAddedPlots,'\n')


    createEXCEL(str(getFilename(FILE_NAME, noext=1)).upper()+' - Plots .xlsx', checkfile=checkfile, autoOpen=0 )
    
    selectedPlots=selectplots(plotsDF)
    createEXCEL(str(getFilename(FILE_NAME, noext=1)).upper()+' - Selection .xlsx', checkfile=0 , selectedPlots=selectedPlots)
    
    return plotsDF


def agriplotcmd():
    Usgstr='''
Usage : Generates Excel file from Plot Data file

        agriplotcmd   FILENAME
        
        Optional : Use option 'checkfile' To Compare with a File
        e.g.  agriplotcmd   FILENAME   --checkfile  FILENAMEToCompare
        
        '''
    argDict=getArg(usagestr=Usgstr, customarg='checkfile')
    if argDict:
        FILE_NAME=argDict['main']
        checkfile=argDict['checkfile']  if argDict.get('checkfile') else ''
        agriplot(FILE_NAME, checkfile=checkfile)


def selectplotscmd():
    Usgstr='''
Usage : Takes 'Total Number of Plots/Last SNo.' and Gives the Selected Plots

        selectplotscmd  Number-of-Plots/Last-SNo.

        '''
    argDict=getArg(usagestr=Usgstr, customarg='checkfile')
    if argDict:
        plotsDF=argDict['main']
        selectedplotdict=selectplots(plotsDF)
        selectedplots=selectedplotdict['splots']
        print('\n\n   Selected Plot SNo.\n\n  '+'\n  '.join([ i for i in selectedplots ]))



def getsubplotscmd():
    Usgstr='''
Usage : Pass the Subplots of a Plot separated by comma in any order to get All the Possible Subplots in Sequence

        getsubplotscmd   Comma-Separated-SubPlots
        '''
    argDict=getArg(usagestr=Usgstr, customarg='checkfile')
    if argDict:
        subplots=argDict['main'].replace('[','').replace(']','')
        splots=getsubplots(subplots)
        print('\n   SubPlots:: \n\n        '+'\n        '.join([ i  for i in splots ])+'\n\n')



if __name__ == '__main__' :

    FILE_NAME='tplots.xlsx'
    #FILE_NAME='plotdata'
    #FILE_NAME='plot1.xlsx'
    #agriplot(FILE_NAME, checkfile='done' )
    agriplot(FILE_NAME )
    
