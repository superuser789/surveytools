
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM, BORDER_THICK
import pandas as pd
from natsort import natsorted
import os, sys
from time import sleep
from tqdm import tqdm
import tkinter as tk
import tkinter.ttk  as ttk
from tkinter import filedialog as fd 
from urllib.request import urlopen
import json, random
import warnings
warnings.filterwarnings('ignore')



verbose=False
GUI=0
#verbose=True
progressmenu=1
startRow=1
plotOne=['-', 0] # SubPlot to Take as 1 
plotsDF = pd.DataFrame(columns=['plotNo','subplot','sno','status']) # DataFrame
n=0 # DataFrame Index
errorMsg=''
VER=1.0

wbg='#174276' 
txtc='#E0E0E0'

def getNumber(cstr): # Get Number at the starting of s string
    num=''
    for c in str(cstr):
      if c.isdigit():
          num+=c
      else:
          break
    return int(num) if num.isdigit() else num

def splitAlpnaNum(sps):
    divlist=[]; tstr=''
    for sp in sps:
        if not sp.isdigit(): # There is alphabet
            divlist.append(tstr) # Add previous numbers
            divlist.append(sp) # Add Current alphabet
            tstr='' # Reset the no.
        else:
            tstr+=sp
    divlist.append(tstr)
    return [ i for i in divlist if i ]
    #import re; return [ i for i in re.split('(\d+)', sps) if i ]


def getRandomNumber(RANCOL, tplot):    
    if hasattr(sys, "_MEIPASS") : # As Pyinstaller Exe
        RandomTable = pd.read_pickle(os.path.join(sys._MEIPASS, 'RandomTable.pkl'))

    elif __name__ == '__main__' : # As Main Script
       #RandomTable = pd.read_pickle( os.path.join( 'data', 'RandomTable.pkl') )
        RandomTable = pd.read_pickle('RandomTable.pkl')
           
    else: # From Command Prompt
        RandomTable = pd.read_pickle(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'RandomTable.pkl'))


    tplotLimit = [ 10**i for i in [1,2,3,4 ] ]
    for tlimit in tplotLimit:
        if tplot <= tlimit:
            rnumlen = len(str(tlimit)) - 1
            break
    rn=0
    for num in RandomTable[RANCOL]:
        rn+=1
        if int(str(num)[:rnumlen]) <= tplot:
            RNUM=int(str(num)[:rnumlen]) if int(str(num)[:rnumlen]) else num
            break
        
    return  RNUM,num,rn


def clear():
    from IPython import get_ipython
    get_ipython().run_line_magic('clear','')
    os.system('cls' if os.name=='nt' else 'clear')


def PRINTUpdateStatus(pstr, addline=''):
    global Txtbox, cwindow, GUI
    print(pstr)
    if GUI:
        if int(radVal.get()) == 1 :
            sleep(0.15)
        Txtbox.insert(tk.END, addline+pstr)
        cwindow.update()

def GapLabel(window,size,nocolor=''):
    GapLabel = tk.Label(window, text = '', font=('calibre', size, 'bold'))
    GapLabel.pack(side = 'top')
    if not nocolor:
        GapLabel.configure(bg=wbg, fg=txtc)

def createTextbox(cwindow):
    global radVal, Txtbox
    windowSize = "550x650+610+10" if int(radVal.get()) < 3 else "400x550+610+10" # Change Window Size for getSubplots
    cwindow.geometry(windowSize)  
    #Txtbox = tk.Text(cwindow,  width = 100, height = 100, bg="#0d3c55", fg="#eed369", font=('calibre', 10, 'bold'))
    Txtbox = tk.Text(cwindow,  width = 100, height = 100, bg="#222233", fg="#aaccff", font=('calibre', 10, 'bold'))
    Txtbox.pack()
    
def update():
    global VER, window, wbg, txtc
    updatelink='https://raw.githubusercontent.com/superuser789/surveytools/main/nsosurveytoolsUpdate.txt'

    UpdateLabel = tk.Label(window, text = '', font=('calibre', 9, 'italic'))
    UpdateLabel.configure(bg=wbg, fg='yellow')
    UpdateLabel.pack(side = 'top', ipadx = 5,  ipady = 10) ; UpdateLabel.place(bordermode=tk.OUTSIDE,  x=10, y=310 )
    
    def updateIt(updateLINK,file):
        if  os.system('start CHROME.EXE "'+updateLINK+'"'):
            os.system('start iexplore.EXE "'+updateLINK+'"')
        sleep(5)
        tk.messagebox.showinfo('Download Update',"DOWNLOAD '"+file+"'.\n and USE it.")
    
    try:
        data = urlopen(updatelink).read().decode('utf-8')
        datadict=json.loads(data)
        version=datadict['ver']
        if float(VER) != float(datadict['ver']):
            UpdateLabel['text'] = 'UPDATE AVAILABLE  !'
            file=[i for i in  datadict['link'].split('\\') if '.exe' in i][0]
            sleep(2)
            tk.messagebox.showinfo('Update v'+datadict['ver']+' Available',"Download Update: '"+file+"'")
            updateIt(datadict['link'],file)
                        
    except Exception as e:
        print('No Internet.', e)
    
    


def Progressbar(window, barname='', UpInterval=5):
    global progressInfo
    GapLabel(window,10)
    style = ttk.Style(window)
    ProgressFrame = tk.Frame(window)
    style.layout('text.Horizontal.TProgressbar',
                 [('Horizontal.Progressbar.trough',
                   {'children': [('Horizontal.Progressbar.pbar',
                                  {'side': 'left', 'sticky': 'ns'})],
                    'sticky': 'nswe'}),
                  ('Horizontal.Progressbar.label', {'sticky': ''})])
                  # , lightcolor=None, bordercolo=None, darkcolor=None
    style.configure('text.Horizontal.TProgressbar', text='0 %')
    
    progress_bar = ttk.Progressbar(ProgressFrame, style='text.Horizontal.TProgressbar', orient="horizontal",mode="determinate", maximum=100, value=0)
     
    progressInfo = tk.Label(ProgressFrame, text=barname ) # Progress Label
     
    # Use the grid manager
    progressInfo.grid(row=0, column=0)
    progress_bar.grid(row=0, column=1)
    ProgressFrame.pack()
    # Necessary, as the master object needs to draw the progressbar widget
    # Otherwise, it will not be visible on the screen
    window.update()
     
    progress_bar['value'] = 0
    window.update()
     
    while progress_bar['value'] < 100:
        progress_bar['value'] += UpInterval
        style.configure('text.Horizontal.TProgressbar', text='{:g} %'.format(progress_bar['value']))
        progressInfo['text']=progress_bar['value']

        window.update() # Keep updating the master object to redraw the progress bar
        sleep(0.5)       



# Gives the Passed Arguments
def getArg(usagestr='', customarg=''): # Gives arguments & Takes Single Custom Arg. optionally
    args = sys.argv[1:] # Exclude Filename from obtained Arguments list
    if len(args) < 1:
        print(usagestr)
        return {}
    else:
        if len(args) > 2:
            if customarg:
                if args[1] == customarg.replace('-','') :
                    return { 'main': args[0], customarg : args[2] }    
        else:
            return { 'main': args[0] }


pval='' ; pltseq=[]
def findMissingPlots(subp): # Takes Subplots of a Plot
    ### Subplots into Dictionary of subplots stored in subplt list
    subplt=[]
    for sps in subp:
        tmpdict=currdict={}
        #spsNo=str(getNumber(sps))
        for sp in splitAlpnaNum(sps) : # Get First 1,2or3 digit no. from subplot
            currdict[sp] = {}
            currdict=currdict[sp]
        subplt.append(tmpdict)

    ### Adding Missing Plots using subplt list
    allplt=[]
    global pval, pltseq
    '''
    def getseq(plt, prevplt):
       extra=[]
       startPlt=65 if ord(plt) > 64 else 49 # startRange for Alphabet & Number
       #LastPlt = 1 if len(prevplt) > 1 else 0
       #if len(prevplt) > 1 :
       extra=['###'+prevplt]
       return [ str(prevplt)+chr(i) for i in range(startPlt, ord(plt)+1)] + extra
    '''
    
    def getseq(plt, prevplt):
       #print(plt,':', prevplt)
       extra=[]
       #if len(prevplt+plt) == 2 : # Converts xA into x
       '''
       if plt=='A': # Add Extra Plots
           plt='B'
       if plt=='1':
           plt='2'
       '''
               
       if plt.isdigit():  # startRange & stopRange for Alphabet & Number
           startPlt=1
           stopPlt=int(plt)
           Intflag=True
       else:
           startPlt=65
           stopPlt=ord(plt)
           Intflag=False
       #startPlt=65 if ord(plt) > 64 else 49 # startRange for Alphabet & Number
       # if len(prevplt+plt) == 2 : # Converts xA into x
       #     if plt=='A':
       #         return [ prevplt, '###'+prevplt+plt ]
       extra=['###'+prevplt] # Remove the Plot Prefix 
       #if len(splitAlpnaNum(prevplt))>2:
       #    return [ str(prevplt)+str(plt)] + extra
       #else:
       return [ str(prevplt)+ ( str(i) if Intflag else chr(i) ) for i in range(startPlt, stopPlt+1)] + extra


  
    def getall(dplt):
        global pval, pltseq
        for k,v in dplt.items():
            #print(k, ',', pval)
            if pval:
                pltseq+=getseq(k, pval)
            pval+=str(k)
            if isinstance(v, dict):
                getall(v)
    
    for sp in subplt: # subplt dict in subplt list
        pval='' ; pltseq=[]
        getall(sp) ; #print(pltseq)
        allplt+=pltseq

    return allplt
    

MissedSinglePlots={}
def getsubplots( subplot, pno=0, addPlot=0): # Gives All Possible subplots of a Plot
    '''
    Pass the List having Subplots in random order to get the Sequenced Subplots
    '''
    global n, plotsDF, plotOne, MissedSinglePlots, GUI, statusLabel
    if not addPlot:
        if not pno:
            plotsDF = pd.DataFrame(columns=['plotNo','subplot','sno','status']); n=0; MissedSinglePlots={}
            
            subplot=[int(i) if str(i).isdigit() else str(i) for i in subplot.split(',')] if isinstance(subplot, str) else subplot
            subplot=[ 1 if i in plotOne else i for i in subplot ]

        #print(subplot)
        for i in subplot:
            if isinstance(getNumber(i),str):
                MSG="INVALID FORMAT !!! \n\n CHECK and CORRECT the Subplot:'"+str(i)+"' of  Plot: '"+str(pno)+"'\n\n  Subplots: "+str(subplot)+' present in EXCEL File \n and RUN Again.'
                if GUI:
                    statusLabel['text']=MSG.replace('\n\n','\n')
                    tk.messagebox.showinfo('INVALID FORMAT !!!', MSG)
                print(MSG)
                    

        #mainRange = max(set([ i if isinstance(i,int) else getNumber(i)  for i in subplot ])) # Get Max. Main Subplots
        mainRange = max(set([ i if isinstance(i,int) else getNumber(i)  for i in subplot ])) # Get Max. Main Subplots
        SinglePlots = [ str(i) for i in subplot if isinstance(i,int) ] # Single Plots into string
        GivenSubPlots=[ i for i in subplot if isinstance(i,str) ] # All SubPlots are string type
        MissedPlots = list(set(findMissingPlots(GivenSubPlots)))
        GivenSingleSubPlots = set([str(getNumber(i)) for i in MissedPlots if i[:3] != '###'])
        if verbose:
            print(str(MissedPlots))
        MissedPlot=MissedPlots.copy()
        ### Remove Extra Plots &  Only single occurrence due to set
        for plt in MissedPlot:
            if '###' == plt[:3]:
                MissedPlots.remove(plt)
                if plt[3:] in MissedPlot:
                    MissedPlots.remove(plt[3:])
                if plt[3:].isdigit() and int(plt[3:]) in MissedPlot: # In case of Integer Plot No.
                    MissedPlots.remove(int(plt[3:]))
        
        [ SinglePlots.remove(i)  for i in GivenSingleSubPlots if i in SinglePlots ] # Remove SinglePlots if it is in GivenSubPlots 

        #AddedPlots = AllPlots - set(GivenSubPlots)
        #AllPlots = sorted( set(SinglePlots) | set(MissedPlots) , key=lambda x: int(''.join(filter(str.isdigit, x))) ) # Combine & Sort
        AllPlots = natsorted( set(SinglePlots) | set(MissedPlots) ) # Combine & Sort
        if verbose:
            print(AllPlots)    
        ## Insert Missing Single Plot if any
        singleplt = list({ getNumber(i) for i in AllPlots }) # All SubPlot Nos. 
        isingleplt = [ i for i in range(1, mainRange+1) ]
        
        if singleplt != isingleplt :
            AddedIntPlots=[ str(i)+'++' for i in isingleplt if i not in singleplt ]
            AllPlots = natsorted( set(SinglePlots+AddedIntPlots) | set(MissedPlots) ) # Combine & Sort    
            if verbose:
                print('\n  Missing In Plot :',pno, ' - ',singleplt, '  Added: ',[i[:-2] for i in AddedIntPlots])
            MissedSinglePlots[pno] = [ singleplt, AddedIntPlots ]
        

        for plt in AllPlots :
            n+=1
            plotsDF.at[n, 'plotNo'] = pno # Add Plot No.
            plotsDF.at[n,'subplot'] = plt
            plotsDF.at[n,'sno'] = n
            
            if plt in SinglePlots : # If Single Plot or Integer
                plotsDF.at[n, 'status'] = 'single'                    
            elif plt in GivenSubPlots:
                plotsDF.at[n, 'status'] = 'division'
            elif '++' == plt[-2:]:
                plotsDF.at[n,'subplot'] = plt[:-2] # overwrite in case of ++ single Plots
                plotsDF.at[n, 'status'] = 'ADDED'
            else:                
                plotsDF.at[n, 'status'] = 'ADDED'                
    
        if verbose:
            print(plotsDF.tail())
    else: # Plot is not Added
        n+=1
        plotsDF.at[n, 'plotNo'] = pno # Add Plot No.
        plotsDF.at[n,'subplot'] = subplot
        plotsDF.at[n,'sno'] = n
        plotsDF.at[n, 'status'] = 'ADDED' 
            
    return  [ i for i in plotsDF['subplot'] ]


def getFilename(FILE_NAME, noext=0):
    FileFormat='.xls'
    if FileFormat in FILE_NAME:
        FileName, extension = str(FILE_NAME).split(FileFormat)
        return FileName if noext else FileName+FileFormat+extension
    else:
        return FILE_NAME if noext else FILE_NAME+'.xlsx'

excelInfo=[]  
def createEXCEL(filename, checkfile=0 , autoOpen=1, selectedPlots=0):
    global plotsDF, startRow, FileFullPath, excelInfo, errorMsg
    
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.title =  'Final'
      
    START_COL='A'
    START_ROW='2'
      
    bold24Font = Font(size=11, bold=True)  # name='Times New Roman'
    bold26Font = Font(size=12, bold=True)  # name='Times New Roman'
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    leftalignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    titleColourTOP='E0E0E0'
    AddedPlotColour='CCFF90'
    
    thick_border = Border(
       left=Side(border_style=BORDER_MEDIUM, color='00000000'),
       right=Side(border_style=BORDER_MEDIUM, color='00000000'),
       top=Side(border_style=BORDER_MEDIUM, color='00000000'),
       bottom=Side(border_style=BORDER_MEDIUM, color='00000000') )
    thin_border = Border(
       left=Side(border_style=BORDER_THIN, color='00000000'),
       right=Side(border_style=BORDER_THIN, color='00000000'),
       top=Side(border_style=BORDER_THIN, color='00000000'),
       bottom=Side(border_style=BORDER_THIN, color='00000000') )
       
    side_border = Border( right=Side(border_style=BORDER_MEDIUM, color='00000000') ) 
    top_border = Border( top=Side(border_style=BORDER_THIN, color='00000000') )
    
    for i in range(3):
        sheet[chr(ord(START_COL) + i)+'1'].font = bold24Font
        sheet[chr(ord(START_COL) + i)+'1'].alignment = alignment  
        sheet[chr(ord(START_COL) + i)+'1'].fill = PatternFill(fgColor=titleColourTOP, fill_type = 'solid')  
        sheet[chr(ord(START_COL) + i)+'1'].border = thick_border
        
    ## Title Names
    sheet[chr(ord(START_COL) ) +str(1)] = 'Survey No.'
    sheet[chr(ord(START_COL)+1 )+str(1)] = 'Sub-Division'
    sheet[chr(ord(START_COL)+2 )+str(1)] = 'S.No.'
    
    AddPltCol='P' ; AddPltRow=4 ; iPlt=3
    if not plotsDF[plotsDF['status']=='ADDED'].empty and selectedPlots :
        ### Added Plots Heading
        cellMergeLen=2
        sheet.merge_cells( AddPltCol+str(1) +':'+ chr(ord(AddPltCol)+cellMergeLen)+str(1) )
        sheet[AddPltCol+str(1)] = 'ADDED PLOTS ('+str(plotsDF['status'][plotsDF['status']=='ADDED'].count())+')'
        sheet[AddPltCol+str(1)].border = thick_border
        sheet[AddPltCol+str(1)].font = bold26Font
        sheet[AddPltCol+str(1)].alignment = alignment
        sheet[AddPltCol+str(1)].fill = PatternFill(fgColor=titleColourTOP, fill_type = 'solid')
        for col in range(16, 16+cellMergeLen + 1): # Apply Style of First Merge cell to the Whole Mergeed cells 
            sheet.cell(row=1, column=col)._style = sheet.cell(row=1, column=16)._style 
        
        sheet[chr(ord(AddPltCol) )  +str(3)] = 'Survey No.'
        sheet[chr(ord(AddPltCol)+1 )+str(3)] = 'Sub-Division'
        sheet[chr(ord(AddPltCol)+2 )+str(3)] = 'S.No.'

        for i in range(3):
            sheet[chr(ord(AddPltCol) + i)+str(iPlt)].font = bold24Font
            sheet[chr(ord(AddPltCol) + i)+str(iPlt)].alignment = alignment  
            sheet[chr(ord(AddPltCol) + i)+str(iPlt)].fill = PatternFill(fgColor=titleColourTOP, fill_type = 'solid')  
            sheet[chr(ord(AddPltCol) + i)+str(iPlt)].border = thick_border


    nextPlotNo=1
    plotInfo={}
    for i in range(len(plotsDF.index)):
        plotInfo.setdefault(plotsDF.iloc[i,0],0)
        plotInfo[plotsDF.iloc[i,0]]+=1
        
        sheet[chr(ord(START_COL) )+str(int(START_ROW)+i)] = plotsDF.iloc[i,0] # plot No.
        sheet[chr(ord(START_COL)+1)+str(int(START_ROW)+i)]= int(plotsDF.iloc[i,1]) if plotsDF.iloc[i,1].isdigit() else plotsDF.iloc[i,1] # SubPlot
        sheet[chr(ord(START_COL)+2)+str(int(START_ROW)+i)]= plotsDF.iloc[i,2] # S.No.


        ## Add Line on Plot No. Change
        if nextPlotNo != plotsDF.iloc[i,0]:
            for j in range(3):
                sheet[chr(ord(START_COL)+j)+str(int(START_ROW)+i)].border = top_border
        nextPlotNo=plotsDF.iloc[i,0]
        
        ### Formatting for Newly Added Subplots
        if plotsDF.iloc[i,3] == 'ADDED' :
            for j in range(3):
                sheet[chr(ord(START_COL)+j)+str(int(START_ROW)+i)].fill = PatternFill(fgColor=AddedPlotColour, fill_type = 'solid')
                sheet[chr(ord(START_COL)+j)+str(int(START_ROW)+i)].border = thin_border
        
        
        ## Cell Formatting
        sheet[chr(ord(START_COL))+str(int(START_ROW)+i)].alignment = alignment
        sheet[chr(ord(START_COL)+1)+str(int(START_ROW)+i)].alignment = leftalignment
        sheet[chr(ord(START_COL)+2)+str(int(START_ROW)+i)].alignment = alignment

    infoStartRow=3 ; infoStartCol='G'
    if selectedPlots:
        sheet.title = 'Selected'
        splots=list(selectedPlots.keys())
        selectedPlots=[i for l in [selectedPlots[i] for i in selectedPlots] for i in l ]
        
        
        ### Frame Selection Heading
        sheet.merge_cells( infoStartCol+str(1) +':'+ chr(ord(infoStartCol)+4)+str(1) )
        sheet[infoStartCol+str(1)] = 'FRAME  SELECTION'
        sheet[infoStartCol+str(1)].border = thick_border
        sheet[infoStartCol+str(1)].font = bold26Font
        sheet[infoStartCol+str(1)].alignment = alignment
        sheet[infoStartCol+str(1)].fill = PatternFill(fgColor=titleColourTOP, fill_type = 'solid')
        for col in range(7, 7+4 + 1): # Apply Style of First Merge cell to the Whole Mergeed cells 
            sheet.cell(row=1, column=col)._style = sheet.cell(row=1, column=7)._style


        
        for i in range(len(plotsDF.index)):
            if plotsDF.iloc[i,2] in selectedPlots:
                pltcolor = '03A9F4' if sheet[chr(ord(START_COL)+2)+str(int(START_ROW)+i)].value in splots else 'B3E5FC'
                sheet[chr(ord(START_COL))+str(int(START_ROW)+i)].fill = PatternFill(fgColor=pltcolor, fill_type = 'solid') # 81D4FA
                
                sheet[chr(ord(START_COL)+1)+str(int(START_ROW)+i)].fill = PatternFill(fgColor=pltcolor, fill_type = 'solid')
                sheet[chr(ord(START_COL))+str(int(START_ROW)+i)].border = thin_border
                sheet[chr(ord(START_COL)+1)+str(int(START_ROW)+i)].border = thin_border

                
            #### Excel Selection Info Formatting
            if  0 <= i-infoStartRow < len(excelInfo) and excelInfo[i-infoStartRow] :
                
                mergeWidth, horzAlign = (5,'left') if ':' in excelInfo[i-infoStartRow] else (2,'left')
                sheet.merge_cells( infoStartCol+str(i) +':'+ chr(ord(infoStartCol)+mergeWidth)+str(i) )
                
                sheet[chr(ord(infoStartCol) )+str(i)] = excelInfo[i-infoStartRow]
                sheet[chr(ord(infoStartCol) )+str(i)].alignment = Alignment(horizontal=horzAlign, vertical='center')
                sheet[chr(ord(infoStartCol) )+str(i)].border = thin_border
                sheet[chr(ord(infoStartCol) )+str(i)].fill = PatternFill(fgColor='E8F5E9', fill_type = 'solid') 

                    
                if '->' in  excelInfo[i-infoStartRow] and i > 10 : # Selected Plots
                    sn,Selplot = excelInfo[i-infoStartRow].replace('.','.      ').replace('[','  ').replace(']',' ').split('.')
                    plt, splt = Selplot.split('->')
                    sheet[chr(ord(infoStartCol) )+str(i)] = '{:>5}'.format(sn) +'.'+ '{:>10}'.format(plt) + '    ->      ' + splt
                    pltcolor = '4FC3F7' if '[' in excelInfo[i-infoStartRow] else 'E1F5FE'
                    sheet[chr(ord(infoStartCol) )+str(i)].fill = PatternFill(fgColor=pltcolor, fill_type = 'solid')
                else:
                    sheet[chr(ord(infoStartCol) )+str(i)].font = bold24Font # Bold All Except Selected Plots
                    if ':' not in excelInfo[i-infoStartRow]: # Selected Plots Heading
                        sheet[chr(ord(infoStartCol) )+str(i)].border = thick_border
                        sheet[chr(ord(infoStartCol) )+str(i)].fill = PatternFill(fgColor='B39DDB', fill_type = 'solid')

                if 'CLUSTER' in excelInfo[i-infoStartRow]: # Cluster Heading
                    sheet[chr(ord(infoStartCol) )+str(i)].fill = PatternFill(fgColor='D1C4E9', fill_type = 'solid')  
                    sheet[chr(ord(infoStartCol) )+str(i)].border = thick_border

                for col in range(7, 7+mergeWidth + 1): # Apply Style of First Merge cell to the Whole Mergeed cells 
                    sheet.cell(row=i, column=col)._style = sheet.cell(row=i, column=7)._style               


            if  plotsDF.iloc[i,3] == 'ADDED' :
                iPlt+=1
                sheet[chr(ord(AddPltCol)  )+str(iPlt)] = plotsDF.iloc[i,0] # plot No.
                sheet[chr(ord(AddPltCol)+1)+str(iPlt)] = int(plotsDF.iloc[i,1]) if plotsDF.iloc[i,1].isdigit() else plotsDF.iloc[i,1] # SubPlot
                sheet[chr(ord(AddPltCol)+2)+str(iPlt)] = plotsDF.iloc[i,2] # S.No.
                
                for ii in range(3):
                    #sheet[chr(ord(AddPltCol) + i)+str(iPlt)].font = bold24Font
                    sheet[chr(ord(AddPltCol) + ii)+str(iPlt)].alignment = alignment  
                    sheet[chr(ord(AddPltCol) + ii)+str(iPlt)].fill = PatternFill(fgColor='CCFF90', fill_type = 'solid')  
                    sheet[chr(ord(AddPltCol) + ii)+str(iPlt)].border = thin_border
                    
                pltcolor = 'B3E5FC' if plotsDF.iloc[i,2] in selectedPlots else 'FFFFFF' 
                sheet[chr(ord(AddPltCol) + 2)+str(iPlt)].fill = PatternFill(fgColor=pltcolor, fill_type = 'solid') 



    else:
        sortedPlots = sorted(plotInfo, key=plotInfo.get, reverse=True) # Based on Subplots count
        '''
        PRINTUpdateStatus('\n\n     Plot No. | Divisions | Probability/%')
        for i,plt in enumerate(sortedPlots[:10]):
            #print(' '+'{:>2}'.format(i+1)+'. '+'{:>5}'.format(plt)+'  '+'{:>9}'.format(plotInfo[plt])+'        '+format((plotInfo[plt]/len(plotsDF.index))*100, '.2f')+' %' )
            PRINTUpdateStatus(' '+'{:>2}'.format(i+1)+'. '+'{:>5}'.format(plt)+'  '+'{:>9}'.format(plotInfo[plt])+'        '+format((plotInfo[plt]/len(plotsDF.index))*100, '.2f')+' %', addline='\n' )
        '''
        if checkfile:
            wbck = openpyxl.load_workbook(getFilename(checkfile))
            sheetck=wbck[wbck.sheetnames[-1]]
            sheetck=wbck.active
            
            ROW_MAX=sheetck.max_row
            
            chkplots={ 'S.No.':'Sub-Division' } # Dictionary to store Plot no. as key & its subplots as value in list form
            for i in range(startRow,ROW_MAX+1):
                #print(pno,' : ',spno)
                pno=sheetck['A'+str(i)].value # plot no.
                if pno:
                    #if str(getNumber(pno)).isdigit():
                    spno = sheetck['B'+str(i)].value
                    sno = sheetck['C'+str(i)].value if sheetck['C'+str(i)].value else i
                    chkplots[sno] = spno
            #cn=int(START_ROW)
            cn=1
            for sno in chkplots:
                sheet['D'+str(cn)] = sno # SNo.
                sheet['E'+str(cn)] = chkplots[sno] # SNo.
                sheet['D'+str(cn)].alignment = alignment; sheet['E'+str(cn)].alignment = alignment
                cn+=1

    try:
        wb.save(FileFullPath+filename)
        wb.close()           
        if autoOpen :
            sleep(4)
            os.system('start EXCEL.EXE "'+FileFullPath+filename+'"')
    except:
        errorMsg = "CLOSE the Opened EXCEL File : \n'"+filename+"'\n and RUN AGAIN."
        PRINTUpdateStatus(errorMsg)
        tk.messagebox.showinfo('Failed to Create File !!!', errorMsg.replace('\n','\n\n'))


def selectplots(plotsDF, args=''):
    global GUI, Txtbox, excelInfo
    print(plotsDF)
    stmgrp=''
    if isinstance(plotsDF,str):
        tplot,stmgrp=tuple(plotsDF.split(',')) if ',' in plotsDF else (plotsDF,'')
    elif isinstance(plotsDF,list):
        tplot,stmgrp=(plotsDF[0],str(plotsDF[1])) if len(plotsDF) == 2 else (plotsDF[0],'')
    else:   
        tplot=len(plotsDF.index)
    tplot=int(tplot)
    def SelectedPlots(i):
        startRange=(i-i%10  if 0<i%10<6 else  i-i%5) if i%10 else i-5
        return [ startRange+i for i in range(1,6) ]

    rtplot = (tplot+abs(tplot%5-5)) if tplot%5 else tplot
    splotinterval=int(rtplot/4)
    stmflag=True; grpflag=True
    print('\n\n')
    
    if not GUI:
        if not stmgrp:
            while(stmflag):
                stm=input('\n  - Enter Stratum : ')
                stmflag=False if stm.isdigit() else True
            while(grpflag):
                grp=input('  - Enter Order of Selection : ')
                grpflag=False if grp.isdigit() else True
            stmgrp=stm+grp 
    else:
         stmgrp=args.replace('-','')
         stm,grp=args.split('-')
    print(stmgrp)
    
    RANCOL = int(stmgrp)%104
    RNUM,oRNUM,rn = getRandomNumber(RANCOL, tplot)

    splots = [ RNUM+splotinterval*i  if RNUM+splotinterval*i <= tplot else (RNUM+splotinterval*i)-rtplot  for i in range(4) ] # Four Randomly Selected Plot SNo.

    PRINTUpdateStatus('\n\n    Random Number : '+str(RNUM)+'  ( In Col. '+str(RANCOL)+(' ('+stmgrp+')' if int(stmgrp) > 104 else ''  )+' at Row '+str(rn)+( ' : '+oRNUM  if RNUM!=int(oRNUM) else '' )+' ) \n\n    H : '+str(tplot)+"    H' : "+str(rtplot)+'\n    Stratum : '+stm+'     Order of Selection : '+grp+'\n    Cluster Interval : '+str(splotinterval))

    excelInfo=[ '    Random Number : '+str(RNUM)+'      ( In Col. '+str(RANCOL)+(' ('+stmgrp+')' if int(stmgrp) > 104 else ''  )+' at Row '+str(rn)+( ' : '+oRNUM  if RNUM!=int(oRNUM) else '' )+' ) ', '    H : '+str(tplot)+"      H' : "+str(rtplot), '    Stratum : '+stm+'       Order of Selection : '+grp , '    Cluster Interval : '+str(splotinterval), '','', '            SELECTED PLOTS', '    SNo.  ->   Plot - Subplot'  ]
    
    selectedplots={} ; 
    if isinstance(plotsDF,pd.core.frame.DataFrame):
        sleep(2)
        PRINTUpdateStatus('\n\n\n    SELECTED PLOTS :: ')
        PRINTUpdateStatus('\n         SNo.    Plot - Subplot\n')
        n=1
        for plt in splots:
            selectedplots[plt]=SelectedPlots(plt)
            for sno in selectedplots[plt]:
                pno=plotsDF[plotsDF['sno']==sno]['plotNo'].values[0]
                subplt=plotsDF[plotsDF['sno']==sno]['subplot'].values[0]
                if not (n-1)%5:
                    excelInfo.append('')
                    excelInfo.append(' '*15+'CLUSTER - '+str(n//5+1))
                if sno==plt:
                    #print(' [','{:>2}'.format(n)+'.','{:>5}'.format(sno),' ->  ',pno,'-',subplt,' ]')
                    excelInfo.append('  ['+'{:>2}'.format(n)+'.'+'{:>5}'.format(sno)+'->'+str(pno)+'-'+str(subplt)+' ]')
                    PRINTUpdateStatus('   ['+'{:>2}'.format(n)+'.'+'{:>5}'.format(sno)+'  ->  '+str(pno)+'-'+str(subplt)+' ]', addline='\n')
                else:
                    #print('  ','{:>2}'.format(n)+'.','{:>5}'.format(sno),' ->  ',pno,'-',subplt)
                    excelInfo.append('    '+'{:>2}'.format(n)+'.'+'{:>5}'.format(sno)+'->'+str(pno)+'-'+str(subplt))
                    PRINTUpdateStatus('    '+'{:>2}'.format(n)+'.'+'{:>5}'.format(sno)+'  ->  '+str(pno)+'-'+str(subplt), addline='\n')
                #print() if not n%5 else ''
                if not n%5:
                    PRINTUpdateStatus('\n\n')
                n+=1
        return selectedplots


    
    else:
        spltstr=[]; n=0
        for plt in splots:
            selectedplots[plt]=SelectedPlots(plt)
            for sno in selectedplots[plt]:
                n+=1
                if sno==plt:
                    spltstr.append('  {:>2}'.format(n)+'. [{:>5}'.format(sno)+']')
                else:
                    spltstr.append('  {:>2}'.format(n)+'.  {:>5}'.format(sno))
                spltstr.append('\n') if not n%5 else ''
        return { 'param':[ RNUM,rtplot,splotinterval ] , 'splots': spltstr }

        

FileFullPath=''
def agriplot(FILE_NAME, checkfile='', cwindow='', args=''):
    global Txtbox
    '''
    Generates Excel file from Plot data file.
    Optionally, Pass the Excel file having manually done Plot sequences in the parameter 'check=' to check with the generated sequences.
    '''
    #print(vars(sys.modules[__name__])['__package__'])
    global FileFullPath
    FILENAME=getFilename(FILE_NAME)
    FileFullPath=os.path.abspath(FILENAME).replace(FILENAME,'')
    wb = openpyxl.load_workbook(getFilename(FILENAME))
    sheet=wb[wb.sheetnames[-1]]
    sheet=wb.active
    global startRow
    ROW_MAX=sheet.max_row
    
    perday=1000
    progressTime=18
    global plotOne
    #ROW_MAX=232
    plots={} # Dictionary to store Plot no. as key & its subplots as value in list form
    for i in range(startRow,ROW_MAX+1):
        #print(pno,' : ',spno)
        pno=sheet['A'+str(i)].value # plot no.
        if pno:
            if str(getNumber(pno)).isdigit(): # If Number in the Starting
                #spno=sheet['B'+str(i)].value # its subplot
                spno = sheet['B'+str(i)].value if sheet['B'+str(i)].value not in plotOne else 1 # its subplot
                plots.setdefault(pno,[])
                plots[pno].append(spno)


    addedPlots=[]
    pn=0; ln=0
    Tplot=sum([ len(plots[pno]) if plots.get(pno) else 1  for pno in range(1,list(plots.keys())[-1]+1) ]) ; ln=0
    LastPlot=list(plots.keys())[-1] # For  Updatebar increment
    progressIncT=progressTime/Tplot # Updatebar wait
    
    if GUI:
        GapLabel(cwindow,10,nocolor=1)
        style = ttk.Style(cwindow)
        ProgressFrame = tk.Frame(cwindow)
        style.layout('text.Horizontal.TProgressbar',
                     [('Horizontal.Progressbar.trough',
                       {'children': [('Horizontal.Progressbar.pbar',
                                      {'side': 'left', 'sticky': 'ns'})],
                        'sticky': 'nswe'}),
                      ('Horizontal.Progressbar.label', {'sticky': ''})])
                      # , lightcolor=None, bordercolo=None, darkcolor=None
        style.configure('text.Horizontal.TProgressbar', text='0 %')
        
        progress_bar = ttk.Progressbar(ProgressFrame, style='text.Horizontal.TProgressbar', orient="horizontal",mode="determinate", maximum=100, value=0)
         
        progressInfoLbl = tk.Label(ProgressFrame, text='0 %', font=('calibre', 10, 'bold')) # Progress Label
         
        # Use the grid manager
        progressInfoLbl.grid(row=0, column=0)
        progress_bar.grid(row=0, column=1)
        ProgressFrame.pack()
        # Necessary, as the master object needs to draw the progressbar widget
        # Otherwise, it will not be visible on the screen
        cwindow.update()
         
        progress_bar['value'] = 0
        cwindow.update()
         

    with tqdm(total=100,bar_format='{desc}{percentage:3.1f}% | {bar}  [{elapsed}<{remaining}]' ) as pbar:    
        for plotNo in range(1,LastPlot+1):
            if plots.get(plotNo):
                pn+=1
                ln+=len(plots[plotNo]) # Show Original Lines for Days calculation
                '''
                    clear()
                    progress = ('.  ' if pn%2 else ' . ' if pn%3 else '  .' )
                    
                    print('\n\n\n      Plot:','{:>3}'.format(plotNo),progress, '{:>5}'.format(int((pn/len(plots))*100)) ,'%  ')
                    print('\n      Days: ', ( 1 if ln <= perday else int(ln/perday)+1), '  |  Lines:','{:>5}'.format(ln) )
                '''
                subplots=getsubplots(plots[plotNo], plotNo)
                #uptL=(len(subplots)/Tplot)*100
            else:
                subplots=getsubplots('1' , plotNo, addPlot=plotNo)
                #ln+=1 # Uncomment to show Added Lines
                addedPlots.append(plotNo)
            progressInfo="  Days:"+str( 1 if ln <= perday else int(ln/perday)+1)+' | Lines:'+str(ln)+' | Plots:'+str(plotNo)+' '
            if progressmenu:
                    pbar.set_description(progressInfo)
                    pbar.update(100/LastPlot)
                    sleep(progressIncT)
            if GUI:
                progress_bar['value'] += 100/LastPlot
                style.configure('text.Horizontal.TProgressbar', text='{:.1f} %'.format(progress_bar['value']))
                progressInfoLbl['text']=progressInfo.replace(':',': ').replace('|',' | ')
                cwindow.update() # Keep updating the master object to redraw the progress bar
                sleep(progressIncT)  

    if GUI:
        createTextbox(cwindow)

    print('\n\n\n')
    if MissedSinglePlots:
        for mplt in MissedSinglePlots:
            #print('\n  Missing In Plot:',mplt, ' - ',MissedSinglePlots[mplt][0], '  Added: ',[i[:-2] for i in MissedSinglePlots[mplt][1]])
            PRINTUpdateStatus('\n  Missing In Plot:'+str(mplt)+' - '+str(MissedSinglePlots[mplt][0])+'  Added: '+str([i[:-2] for i in MissedSinglePlots[mplt][1]]))
    if addedPlots:
        PRINTUpdateStatus('\n\n  Added Single Plots:'+str(addedPlots))
        
    TotalAdded=plotsDF[plotsDF['status']=='ADDED']['status']
    if not TotalAdded.empty:
        TotalAddedPlots=int(TotalAdded.value_counts())
        PRINTUpdateStatus('\n\n  TOTAL Plots/SubPlots[Lines] Added: '+str(TotalAddedPlots)+'\n')

        #print('  TOTAL Plots/SubPlots : ',len(plotsDF.index),' = ',ln,'+',TotalAddedPlots,'\n')


    createEXCEL(str(getFilename(FILE_NAME, noext=1)).upper()+' - Plots .xlsx', checkfile=checkfile, autoOpen=0 )
    
    selectedPlots=selectplots(plotsDF, args=args)
    createEXCEL(str(getFilename(FILE_NAME, noext=1)).upper()+' - Selection .xlsx', checkfile=0 , selectedPlots=selectedPlots)
    
    return plotsDF


def agriplotcmd():
    Usgstr='''
Usage : Generates Excel file from Plot Data file

        agriplotcmd   FILENAME including its FullPath if the file is not in the current directory
        
        e.g.  agriplotcmd   plots.xlsx
        
        Optional : Use option 'checkfile' To Compare with a File
        e.g.  agriplotcmd   FILENAME   --checkfile  FILENAMEToCompare
        
        '''
    argDict=getArg(usagestr=Usgstr, customarg='checkfile')
    if argDict:
        FILE_NAME=argDict['main']
        checkfile=argDict['checkfile']  if argDict.get('checkfile') else ''
        agriplot(FILE_NAME, checkfile=checkfile)



def selectplotsFile(FILENAME, cwindow='', args=''):
    global plotsDF, startRow, FileFullPath, GUI, Txtbox
    FileFullPath=os.path.abspath(FILENAME).replace(FILENAME,'')
    #PRINTUpdateStatus('\n  - Reading File: '+FILENAME)
    wb = openpyxl.load_workbook(FILENAME)
    sheet=wb[wb.sheetnames[-1]]
    sheet=wb.active
    ROW_MAX=sheet.max_row


    if GUI:
        progressTime=0.001; perday=1000
        progressIncT=progressTime/ROW_MAX # Updatebar wait
        GapLabel(cwindow,10,nocolor=1)
        style = ttk.Style(cwindow)
        ProgressFrame = tk.Frame(cwindow)
        style.layout('text.Horizontal.TProgressbar',
                     [('Horizontal.Progressbar.trough',
                       {'children': [('Horizontal.Progressbar.pbar',
                                      {'side': 'left', 'sticky': 'ns'})],
                        'sticky': 'nswe'}),
                      ('Horizontal.Progressbar.label', {'sticky': ''})])
                      # , lightcolor=None, bordercolo=None, darkcolor=None
        style.configure('text.Horizontal.TProgressbar', text='0 %')
        
        progress_bar = ttk.Progressbar(ProgressFrame, style='text.Horizontal.TProgressbar', orient="horizontal",mode="determinate", maximum=100, value=0)
         
        progressInfoLbl = tk.Label(ProgressFrame, text='0 %', font=('calibre', 10, 'bold')) # Progress Label
         
        # Use the grid manager
        progressInfoLbl.grid(row=0, column=0)
        progress_bar.grid(row=0, column=1)
        ProgressFrame.pack()
        # Necessary, as the master object needs to draw the progressbar widget
        # Otherwise, it will not be visible on the screen
        cwindow.update()
         
        progress_bar['value'] = 0
        cwindow.update()


    ### Reading Excel File
    plotsDF = pd.DataFrame(columns=['plotNo','subplot','sno','status']) ; n=0
    for i in range(startRow,ROW_MAX+1):
        pno=sheet['A'+str(i)].value # plot no.
        if pno:
            if str(getNumber(pno)).isdigit(): # If Number in the Starting  
                n+=1
                plotsDF.at[n, 'plotNo']  = sheet['A'+str(i)].value 
                plotsDF.at[n, 'subplot'] = str(sheet['B'+str(i)].value)
                plotsDF.at[n, 'sno']     = n
                plotsDF.at[n, 'status']  = ' '
        if GUI:
            progressInfo="  Days:"+str( 1 if n <= perday else int(n/perday)+1)+' | Lines:'+str(n)+' | Plots:'+str(sheet['A'+str(i)].value)+' '
            progress_bar['value'] += 100/ROW_MAX
            style.configure('text.Horizontal.TProgressbar', text='{:.1f} %'.format(progress_bar['value']))
            progressInfoLbl['text']=progressInfo.replace(':',': ').replace('|',' | ')
            cwindow.update() # Keep updating the master object to redraw the progress bar
            #sleep(progressIncT)              
    if GUI:
        createTextbox(cwindow)               
                
                
    selectedplotdict=selectplots(plotsDF, args=args)
    sleep(2)
    createEXCEL(str(getFilename(FILENAME, noext=1)).upper()+' - Selection .xlsx', checkfile=0 , selectedPlots=selectedplotdict)
    return selectedplotdict

            

def selectplotscmd():
    global FileFullPath, startRow, plotsDF
    Usgstr='''
Usage : Takes 'Total Number of Plots/Last SNo.' and Gives the Selected Plots
                        OR
        Takes FileName and Marks the Selected Plots 

        selectplotscmd  Number-of-Plots/Last-SNo OR FILENAME
        
        e.g.  selectplotscmd  9788
              selectplotscmd  FILENAME

        '''
    argDict=getArg(usagestr=Usgstr)
    if argDict:
        plotsDF=argDict['main']
        if plotsDF.isdigit():
            selectedplotdict=selectplots(plotsDF)
            selectedplots=selectedplotdict['splots']
            print('\n\n   Selected Plot SNo.\n\n  '+'\n  '.join([ i for i in selectedplots ]))
        else:
            FILENAME=getFilename(argDict['main'])
            selectplotsFile(FILENAME)



def getsubplotscmd():
    Usgstr='''
Usage : Pass the Subplots of a Plot separated by comma in any order to get All the Possible Subplots in Sequence

        getsubplotscmd   Comma-Separated-SubPlots
        
        e.g.  getsubplotscmd  6A,6,2C3B,10,7A1,4D5C3B1
        '''
    argDict=getArg(usagestr=Usgstr, customarg='checkfile')
    if argDict:
        subplots=argDict['main'].replace('[','').replace(']','')
        splots=getsubplots(subplots)
        ppt=''; splot=[]
        for splt in splots:
            if getNumber(splt) != ppt:
                ppt=getNumber(splt)
                splt='\n        '+splt
            splot.append(splt)
        print('\n   SubPlots:: \n\n      '+'\n      '.join([ '- '+i.replace('\n','').strip()  if  i.replace('\n','').strip() in subplots.split(',') else '  '+i for i in splot ])+'\n\n')


# progressInfo=''
window=''; cwindow=''; inputFrame=''; Txtbox='' ; name_entry='' ; radVal=''; statusLabel=''
singleVar=''; stratumVar=''; osVar=''
def agriculturesurvey():
    global wbg, txtc, inputFrame, Txtbox, progressInfo, radVal, statusLabel
    global singleVar, stratumVar, osVar, GUI
    
    
    GUI=1
    
    window = tk.Tk()
    window.configure(bg=wbg)
    window.title("Agriculture Survey  v"+str(VER))
    window.geometry("450x360+150+150") # Width x Height
    
    singleVar=tk.StringVar(window,'') 
    stratumVar=tk.StringVar(window,'') 
    osVar=tk.StringVar(window,'')
    
    separator = ttk.Separator(window, orient='horizontal')
    inputFrame = tk.Frame(window)
    
    #radVal = tk.StringVar(window, "1")  # Tkinter string variable 
    radVal = tk.StringVar(window, '1')  # Tkinter string variable 
    prg = tk.IntVar(window, 1)
    checkFile = tk.IntVar()

    
    programs = {" Input Excel File   " : "1", 
                " Selection          " : "2", 
                " Check SubDivisions " : "3"  }
    programInfo = { '1': 'Finds & Adds the Missing Plots with Frame Selection' ,
                    '2': 'Only Plots Selection ',
                    '3': 'Gives All the Possible SubDivisions of a Plot in Sequence' }

    
    def RunProg():
        global window, cwindow, singleVar, name_entry, stratumVar, osVar, Txtbox, errorMsg, plotsDF, n
        
        plotsDF = pd.DataFrame(columns=['plotNo','subplot','sno','status']); n=0
        
        currentProgram = list(programs.keys())[int(radVal.get())-1]
        print('Running ...',list(programs.keys())[int(radVal.get())-1])
        statusLabel['text'] = 'WAIT..  Running... '+currentProgram
        
        cwindow = tk.Toplevel(window)
        cwindow.title("Agriculture Survey : "+currentProgram)

        windowSize = "550x80+610+10" if int(radVal.get()) < 3 else "400x550+610+10" # Change Window Size for getSubplots

        cwindow.geometry(windowSize)
        #cwindow.geometry()
        ### Add Text Box to show Program Status
        
        # if int(radVal.get()) > 1 : 
        #     Txtbox = tk.Text(cwindow,  width = 400, height = 500)
        #     Txtbox.pack()
        
        if int(radVal.get()) < 3 :
            InputFileData=singleVar.get()
            print('ENTERED:',InputFileData)
            FILENAME=getFilename(InputFileData)
            File=os.path.abspath(FILENAME)        
            isFILE=os.path.isfile(File)
            
            NotEnteredFields='Enter '
            mbx=False #if isFILE or InputFileData.isdigit() else True
            
            if not InputFileData:
                NotEnteredFields+= "'Excel File' " if int(radVal.get())==1 else "'Excel File or Total Plots' " ; mbx=True
            if not stratumVar.get():
                NotEnteredFields+="'Stratum' " ; mbx=True
            if not osVar.get():
                NotEnteredFields+="'Order of Selection' " ; mbx=True
    
            TitleMsg, BxMsg = ( 'No Data',NotEnteredFields+'.'  )
            #TitleMsg, BxMsg = ( 'Not Found !!!',"  File Not Found : '"+File+"'" )  if not isFILE and not InputFileData.isdigit() else   ( 'No Data',NotEnteredFields+'.')

            if mbx: 
                cwindow.destroy()
                tk.messagebox.showinfo(TitleMsg, BxMsg)
                statusLabel['text'] = programInfo[radVal.get()]
                
            elif int(radVal.get()) == 1 :
                checkfile = File if checkFile.get() else ''
                
                #t1 = threading.Thread( target = agriplot, args=(File,), kwargs={ 'checkfile':checkfile, 'cwindow':cwindow, 'args':stratumVar.get()+osVar.get() } ) ; t1.start()
                plotsdf=''
                plotsdf = agriplot(File, checkfile=checkfile, cwindow=cwindow, args=stratumVar.get()+'-'+osVar.get())
                if isinstance(plotsdf, str):
                    statusLabel['text'] = 'Something went Wrong.'
                else:
                    statusLabel['text'] = "Generated : \n'"+str(getFilename(InputFileData, noext=1)).upper()+"-Plots'\n & '"+str(getFilename(InputFileData, noext=1)).upper()+"-Selection'  Files."
            elif int(radVal.get()) == 2 :
                selectedplotdict='' ; TPlotLIMIT = 10**4
                if InputFileData.isdigit():
                    if int(InputFileData) <= TPlotLIMIT :
                        createTextbox(cwindow)
                        #Txtbox = tk.Text(cwindow,  width = 400, height = 500) ; Txtbox.pack()
                        selectedplotdict=selectplots(InputFileData, args=stratumVar.get()+'-'+osVar.get())
                        selectedplots=selectedplotdict['splots']
                        Txtbox.insert(tk.END,'\n\n\n   Selected Plot SNo.\n\n  '+'\n  '.join([ i for i in selectedplots ]))                 
                    else:
                        statusLabel['text'] = " H can't exceed "+str(TPlotLIMIT)+'.' 
                        cwindow.destroy()
                        tk.messagebox.showinfo('H Limit !!!'," H can't exceed "+str(TPlotLIMIT)+'.')

                else:
                    selectedplotdict=selectplotsFile(File, cwindow=cwindow, args=stratumVar.get()+'-'+osVar.get())

                if isinstance(selectedplotdict, dict):
                    statusLabel['text'] = 'Selection Done.' if InputFileData.isdigit() else "Selected Plots marked in \n '"+str(getFilename(InputFileData, noext=1)).upper()+"-Selection'  File."
                else:
                    statusLabel['text'] = 'Something went Wrong.'                    
                    

        elif int(radVal.get()) == 3 :
            InputFileData=name_entry.get(1.0, "end-1c")
            SPACE='' ; splots=''
            if InputFileData:
                createTextbox(cwindow)
                #Txtbox = tk.Text(cwindow,  width = 400, height = 500); Txtbox.pack()
                
                subplots=InputFileData.replace('[','').replace(']','')
                subplots=''.join([ i.upper() if i.isalpha() else i for i in subplots  ])
                splots=getsubplots( subplots )
                ppt=''; splot=[]
                for splt in splots:
                    if getNumber(splt) != ppt:
                        ppt=getNumber(splt)
                        splt='\n        '+splt
                    splot.append(splt)
                SUBPLOTS='\n  Sub Divisions for  [ '+', '.join([ i.replace(' ','')  for i in subplots.split(',')])+' ] : \n\n\n      '+'\n      '.join([ '- '+i.replace('\n','').strip()  if  i.replace('\n','').strip() in subplots.split(',') else '  '+i for i in splot ])+'\n\n'
                Txtbox.insert(tk.END, SUBPLOTS)
                statusLabel['text'] = 'Sub Divisions Generated.' if isinstance(splots, list)  else 'Something went Wrong.' 
                
            else:
                cwindow.destroy()
                tk.messagebox.showinfo('No Data !!!',"  Enter Subplots of a Plot.\n\n  e.g.  6A,6,5,2C3B,10,7A1,4D5C3B,8C2")
                statusLabel['text'] = programInfo[radVal.get()]


        if errorMsg:
            statusLabel['text'] = errorMsg
            Txtbox.delete('1.0', tk.END)
            Txtbox.insert(tk.END, ' - '+errorMsg.replace('\n',''))
        else:
            windowSize = "550x650+610+10" if int(radVal.get()) < 3 else "400x550+610+10" # Change Window Size for getSubplots
        Txtbox.configure(state="disabled")
        cwindow.geometry(windowSize)    


    def radioSelection():
        selection = "You selected the option " + str(radVal.get())
        currentProgram = list(programs.keys())[int(radVal.get())-1]
        print('Selected:',radVal.get(),currentProgram)
        #statusLabel['text'] = defaultLabel+currentProgram
        statusLabel['text'] = programInfo[radVal.get()]
        enterData()
    
    def browsefile():
        filename = fd.askopenfilename()
        name_entry.delete(0, tk.END)
        name_entry.insert(0, filename)
    
    
    def enterData():
        global inputFrame, singleVar, stratumVar, osVar, name_entry
        if int(radVal.get()):
            inputFrame.destroy()
            #inputFrame.grid_forget()
        inputFrame = tk.Frame(window)
        inputFrame.configure(background=wbg)
        name_entry_row=0; name_entry_col=1 # Excel FileName
        if int(radVal.get()) < 3:
            mainlabel = 'Excel File :' if int(radVal.get()) == 1 else 'ExcelFile or H :'
            name_entry = tk.Entry(inputFrame, width=30, textvariable=singleVar,font=('calibre',10,'bold'), bg=txtc)
        else:
            mainlabel = 'Enter SubDivisions of a Plot :' 
            name_entry = tk.Text(inputFrame, height = 6, width = 20, bg=txtc, font=('calibre',11,'bold'))
            name_entry_row=1; name_entry_col=0 # Excel FileName
            
        name_label = tk.Label(inputFrame, text = mainlabel, font=('calibre', 10, 'bold'))    
        name_label.grid(row=0,column=0) 
        name_entry.grid(row=name_entry_row,column=name_entry_col)
        name_label.configure(bg=wbg, fg=txtc)

        ## Entry for Stratum, OS
        if int(radVal.get()) < 3 :
            #browsebutton =  tk.Button(inputFrame, text="Browse", command=browsefile,  width=7, bg="#E1F5FE", fg="black", font=('calibre', 8, 'bold'))
            browsebutton =  tk.Button(inputFrame, text="Browse", command=browsefile,  width=7, bg="#add8ff", fg="black", font=('calibre', 8, 'bold'))
            stratum_label = tk.Label(inputFrame, text = 'Enter Stratum :', font=('calibre', 10, 'bold'))   
            stratum_entry = tk.Entry(inputFrame, width=10, textvariable=stratumVar,font=('calibre',10,'bold'), bg=txtc) 
            
            #stratum_entry.tag_configure("bold-and-red", font=Font(size=12, weight="bold"), background="red")
            
            
            os_label = tk.Label(inputFrame, text = 'Order of Selection :', font=('calibre', 10, 'bold'))   
            os_entry = tk.Entry(inputFrame,  width=10, textvariable=osVar,font=('calibre',10,'bold'), bg=txtc)
            if int(radVal.get()) == 1 :
                chkFileButton = tk.Checkbutton(inputFrame, text = "Compare ?", variable = checkFile, onvalue = 1, offvalue = 0, height = 2, width = 15, selectcolor=wbg, bg=wbg, fg=txtc, font=('calibre', 9, 'bold'))
                #chkFileButton = tk.Checkbutton(inputFrame, variable=checkFile, onvalue=1, offvalue=0, text="Compare ?", bg=wbg, fg=txtc, activebackground='black', activeforeground='black',selectcolor=wbg, height = 2, width = 15)
                chkFileButton.grid(row=3,column=1)
                
    
            browsebutton.grid(row=0,column=2) 
            stratum_label.grid(row=1,column=0) 
            stratum_entry.grid(row=1,column=1)
            os_label.grid(row=2,column=0) 
            os_entry.grid(row=2,column=1)
            stratum_label.configure(bg=wbg, fg=txtc)
            os_label.configure(bg=wbg, fg=txtc)
        inputFrame.pack()
    
    
    defaultLabel='Selected : '
    programSelection=tk.StringVar()
    programSelection.set(defaultLabel)
    #label = tk.Label(window, text="Selected Program : ", textvariable=programSelection).pack(side = 'top',  ipady = 5)
    
    GapLabel(window,4) # After Radio Button
    #statusLabel = tk.Label(window, text = defaultLabel+list(programs.keys())[0])
    statusLabel = tk.Label(window, text = programInfo[radVal.get()], font=('calibre', 9, 'italic'))
    statusLabel.configure(bg=wbg, fg=txtc)
    statusLabel.pack(side = 'top',  ipady = 5)
    #label1.place(x=0, y=0)
    

    style = ttk.Style(window) 
    style.configure("TRadiobutton",  background = wbg,  foreground = txtc,
                    font = ("arial", 11, "bold")) 
    ## Radio Buttons
    RadioFrame = tk.Frame(window); RadioFrame.configure(background=wbg)
    for (name, value) in programs.items(): 
        #tk.Radiobutton(window, text = name, variable = programSel, value = value).pack() 
        radbutton=ttk.Radiobutton(RadioFrame, command=radioSelection, text = name, variable = radVal, value = value); radbutton.pack(side = 'top', anchor = 'w',  ipady = 5) ;# radbutton.configure(selectcolor=wbg )
    RadioFrame.pack()    
    
    
    ## Line Separator
    separator.pack(side='top', fill='x')
    
    
    
    ## Run Button
    #runButton = tk.Button( text="START", command=RunProg, width=10, height=1, bg="#80DEEA", fg="black", font=('calibre', 10, 'bold') )
    runButton = tk.Button( text="START", command=RunProg, width=10, height=1, bg="#80b9ee", fg="black", font=('calibre', 10, 'bold') )
    runButton.pack(side = 'top', ipadx = 5,  ipady = 10)
    runButton.place(bordermode=tk.OUTSIDE,  x=180, y=305 )
    
    
    GapLabel(window,10)
    enterData()
    if random.randint(0,100)%3:
        print('Checking for Update...')
        update()
    window.mainloop()    
    


if __name__ == '__main__' :

    FILE_NAME='plots.xlsx'
    #FILE_NAME='plotdata'
    #agriplot(FILE_NAME, checkfile='done' )
    #agriplot(FILE_NAME )
    agriculturesurvey()
    
